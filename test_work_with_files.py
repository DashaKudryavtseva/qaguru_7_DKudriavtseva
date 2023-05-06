import csv
from zipfile import ZipFile
import pytest
from openpyxl import load_workbook
from PyPDF2 import PdfReader


@pytest.fixture
def archiving_files():
    files_for_pack = ["file_example_XLSX_10.xlsx", "sample.pdf", "username.csv"]
    with ZipFile("Simple.zip", mode="w") as zip_file:
        for file in files_for_pack:
            zip_file.write(file)


def xlsx_content_check(file):
    xl_file = load_workbook((file))
    s = xl_file.active
    name_s = xl_file.sheetnames
    assert len(name_s) == 1
    assert name_s[0] == "Sheet1"
    assert s.max_row == 10
    assert s.max_column == 8
    xl_file.close()


def test_check_zip_xlsx():
    with ZipFile("Simple.zip") as check_zip:
        name_files = check_zip.namelist()
        for file in name_files:
            if ".xlsx" in file:
                assert file == "file_example_XLSX_10.xlsx"
                xlsx_content_check(file)


def csv_content_check(file):
    with open(file, "r") as csv_file:
        users = csv.DictReader(csv_file, delimiter=";")
        users = [user for user in users]
        assert len(users) == 5
        assert users[0]["Username"] == "booker12"


def test_check_zip_csv():
    with ZipFile("Simple.zip") as check_zip:
        name_files = check_zip.namelist()
        for file in name_files:
            if ".csv" in file:
                assert file == "username.csv"
                csv_content_check(file)


def pdf_content_check(file):
    reader = PdfReader(file)
    assert len(reader.pages) == 2
    index = reader.pages[0].extract_text().find('This is a small demonstration .pdf file')
    assert index == 21

def test_check_zip_pdf():
    with ZipFile("Simple.zip") as check_zip:
        name_files = check_zip.namelist()
        for file in name_files:
            if ".pdf" in file:
                assert file == "sample.pdf"
                pdf_content_check(file)
